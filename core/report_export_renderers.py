#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""HTML, PDF, XLSX, figure and checksum renderers for project reports."""
from __future__ import annotations

import html
import logging
import uuid
from pathlib import Path
from typing import Any

import numpy as np

from core.gis_map_export import export_project_plan_map
from core.plot_font_policy import configure_matplotlib_cjk_fonts
from core.report_export_rows import _sha256_file

_LOGGER = logging.getLogger(__name__)
from core.tabular_security import safe_tabular_value

def _write_xlsx_report(
    path: Path,
    *,
    summary: dict[str, Any],
    sheets: list[tuple[str, list[dict[str, Any]]]],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError("生成 Excel 工程汇总需要 openpyxl。") from exc
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    project = summary.get("project", {})
    overview = wb.create_sheet("项目概况")
    overview.append(["MyGPR 工程成果汇总", ""])
    overview.merge_cells("A1:B1")
    overview["A1"].font = Font(bold=True, size=16)
    overview["A1"].alignment = Alignment(horizontal="center")
    for key, label in [
        ("name", "项目名称"), ("project_no", "项目编号"), ("location", "测区位置"),
        ("operator", "编制/操作员"), ("reviewer", "复核人"), ("approver", "批准人"),
        ("device_model", "设备型号"), ("coordinate_system", "坐标系统"),
        ("vertical_datum", "高程基准"), ("generated_at", "生成时间"),
    ]:
        value = summary.get("generated_at", "") if key == "generated_at" else project.get(key, "")
        overview.append([label, safe_tabular_value(value)])
    overview.column_dimensions["A"].width = 22
    overview.column_dimensions["B"].width = 72
    overview.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for name, rows in sheets:
        ws = wb.create_sheet(name[:31])
        headers = list(rows[0].keys()) if rows else ["说明"]
        ws.append(headers)
        if not rows:
            ws.append(["无记录"])
        else:
            for row in rows:
                ws.append([safe_tabular_value(row.get(header, "")) for header in headers])
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx, header in enumerate(headers, start=1):
            values = [str(header)] + [str(row.get(header, "")) for row in rows[:200]]
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max(len(v) for v in values) + 2, 10), 42)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp.xlsx")
    wb.save(tmp)
    tmp.replace(path)

def _write_report_figures(store: Any, figures_dir: Path, *, line_ids: set[str] | None = None, include_plan_map: bool = True, cancel_checker=None, progress_callback=None) -> list[Path]:
    import matplotlib.pyplot as plt
    figures_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[Path] = []
    lines = [line for line in store.list_lines() if line_ids is None or line.line_id in line_ids]

    # Rebuild the plan map from registered GIS/trajectory/interface data instead
    # of exporting the interactive Qt preview or a trajectory-only sketch.
    plan_path = figures_dir / "project_plan_map.png"
    try:
        if not include_plan_map:
            raise LookupError("plan map excluded by report profile")
        export_project_plan_map(
            store,
            plan_path,
            cancel_requested=cancel_checker,
            progress_callback=(
                (lambda current, total, message: progress_callback(current, max(total, 1), message))
                if progress_callback is not None else None
            ),
            dpi=180,
        )
        outputs.append(plan_path)
    except Exception as exc:
        from core.job_manager import JobCancelled
        if isinstance(exc, JobCancelled):
            raise
        if isinstance(exc, LookupError):
            pass
        else:
            # Projects without real spatial coordinates may still produce valid
            # B-scan/interface figures and a formal report with an explicit gap.
            store.append_log(f"报告 GIS 平面图未生成: {exc}")

    for index, line in enumerate(lines, start=1):
        if cancel_checker is not None and cancel_checker():
            from core.job_manager import JobCancelled
            raise JobCancelled("报告图件生成已取消")
        try:
            dataset=store.load_gpr_dataset(line.line_id)
        except Exception:
            continue
        matrix=dataset.normalized_preview(max_samples=900,max_traces=1800)
        row_step=max(1,int(np.ceil(dataset.sample_count/matrix.shape[0]))); col_step=max(1,int(np.ceil(dataset.trace_count/matrix.shape[1])))
        distance=np.asarray(dataset.distance_axis_m[::col_step][:matrix.shape[1]]); depth=np.asarray(dataset.depth_axis_m[::row_step][:matrix.shape[0]])
        configure_matplotlib_cjk_fonts()
        fig,ax=plt.subplots(figsize=(10.5,4.2),dpi=150)
        ax.imshow(matrix,cmap="gray",aspect="auto",origin="upper",extent=[float(distance[0]),float(distance[-1]),float(depth[-1]),float(depth[0])],vmin=-1,vmax=1)
        try:
            ann=store.load_basal_interface_annotation(line.line_id)
            if ann is not None:
                curve=ann.curve_samples(); finite=np.isfinite(curve)
                if finite.any():
                    y=np.interp(curve[finite],np.arange(dataset.sample_count),dataset.depth_axis_m)
                    ax.plot(np.asarray(dataset.distance_axis_m)[finite],y,linewidth=1.7,label="基覆界面")
                    ax.legend(fontsize=8)
        except Exception:  # noqa: BLE001 - 注解缺失时图件继续生成，仅告警
            _LOGGER.warning("基覆界面注解加载失败（图件继续生成）：%s", line.line_id)
        ax.set_title(f"{line.line_id} 原始 B-scan 与基覆界面")
        ax.set_xlabel("距离 (m)"); ax.set_ylabel("深度 (m)")
        out=figures_dir/f"{line.line_id}_basal_interface.png"; fig.tight_layout(); fig.savefig(out,bbox_inches="tight"); plt.close(fig); outputs.append(out)
        if progress_callback is not None:
            progress_callback(index, max(len(lines),1), f"生成图件 {line.line_id}")
    return outputs

def _write_checksums(package_dir: Path, path: Path, *, cancel_checker=None) -> list[dict[str, Any]]:
    rows=[]
    for file in sorted(p for p in package_dir.rglob("*") if p.is_file() and p != path):
        rows.append({
            "path": file.relative_to(package_dir).as_posix(),
            "size_bytes": file.stat().st_size,
            "sha256": _sha256_file(file, cancel_checker=cancel_checker),
        })
    text="".join(f"{row['sha256']}  {row['path']}\n" for row in rows)
    path.write_text(text,encoding="utf-8")
    return rows

def _write_html_report(path: Path, *, summary: dict[str, Any], lines: list[dict[str, Any]], quality: list[dict[str, Any]], interfaces: list[dict[str, Any]], targets: list[dict[str, Any]], artifacts: list[dict[str, Any]], spatial: list[dict[str, Any]], sensor_sync: list[dict[str, Any]] | None = None, gis_layers: list[dict[str, Any]] | None = None, boreholes: list[dict[str, Any]] | None = None, figure_paths: list[Path] | None = None) -> None:
    def table(headers: list[str], rows: list[dict[str, Any]], limit: int = 30) -> str:
        head = "".join(f"<th>{html.escape(h)}</th>" for h in headers)
        body_rows = []
        for row in rows[:limit]:
            body_rows.append("<tr>" + "".join(f"<td>{html.escape(str(row.get(h, '')))}</td>" for h in headers) + "</tr>")
        if not body_rows:
            body_rows.append(f"<tr><td colspan='{len(headers)}'>无记录</td></tr>")
        return f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"

    project = summary.get("project", {})
    metrics = summary.get("metrics", {})
    sensor_sync = sensor_sync or []
    gis_layers = gis_layers or []
    boreholes = boreholes or []
    figure_paths = figure_paths or []
    generated_at = summary.get("generated_at", "")
    snapshot = dict(summary.get("snapshot") or {})
    source_binding = dict(summary.get("source_binding") or {})
    lifecycle = dict(summary.get("lifecycle") or {})
    css = """
    body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Microsoft YaHei', sans-serif; margin: 28px; color: #1f2937; }
    h1 { margin-bottom: 4px; } h2 { margin-top: 28px; border-bottom: 1px solid #d1d5db; padding-bottom: 6px; }
    .meta, .metric { color: #4b5563; } .metrics { display: flex; flex-wrap: wrap; gap: 10px; margin: 18px 0; }
    .metric { border: 1px solid #d1d5db; border-radius: 8px; padding: 10px 12px; min-width: 130px; background: #f9fafb; }
    .metric b { display: block; font-size: 22px; color: #111827; }
    .approval { display:grid; grid-template-columns: repeat(3, 1fr); gap:12px; margin:18px 0; }
    .approval div { border:1px solid #cbd5e1; min-height:72px; padding:10px; }
    figure { margin:18px 0; page-break-inside:avoid; } figure img { max-width:100%; border:1px solid #d1d5db; }
    table { border-collapse: collapse; width: 100%; margin-top: 10px; font-size: 12px; }
    th, td { border: 1px solid #d1d5db; padding: 6px 7px; vertical-align: top; }
    th { background: #eef2ff; text-align: left; } code { background:#f3f4f6; padding:2px 4px; border-radius:4px; }
    """
    html_text = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><title>{html.escape(project.get('name', 'MyGPR 项目报告'))}</title><style>{css}</style></head>
<body>
<h1>{html.escape(project.get('name', 'MyGPR 项目'))} 成果报告</h1>
<div class="meta">生成时间：{html.escape(str(generated_at))}　项目编号：{html.escape(str(project.get('project_no', '--')))}　测区：{html.escape(str(project.get('location', '--')))}</div>
<div class="meta">快照：{html.escape(str(snapshot.get('snapshot_id', '--')))}　项目修订：{html.escape(str(snapshot.get('project_revision', '--')))}　空间成果：{html.escape(str(source_binding.get('spatial_result_id', '未绑定') or '未绑定'))}　审签状态：{html.escape(str(lifecycle.get('approval_status', '工作草稿')))}</div>
<div class="metrics">
  <div class="metric"><b>{metrics.get('line_count', 0)}</b>测线</div>
  <div class="metric"><b>{metrics.get('imported_line_count', 0)}</b>已导入测线</div>
  <div class="metric"><b>{metrics.get('qc_passed_count', 0)}</b>质检通过</div>
  <div class="metric"><b>{metrics.get('processing_artifact_count', 0)}</b>处理结果</div>
  <div class="metric"><b>{metrics.get('interface_line_count', 0)}</b>界面标注测线</div>
  <div class="metric"><b>{metrics.get('spatial_export_count', 0)}</b>空间成果</div>
  <div class="metric"><b>{metrics.get('borehole_passed_count', 0)}/{metrics.get('borehole_comparison_count', 0)}</b>钻孔误差达标</div>
</div>
<h2>签署与版本</h2>
<div class="approval"><div><b>编制</b><br>{html.escape(str(project.get('compiler') or project.get('operator') or '--'))}<br><br>签字/日期：</div><div><b>复核</b><br>{html.escape(str(project.get('reviewer') or '--'))}<br><br>签字/日期：</div><div><b>批准</b><br>{html.escape(str(project.get('approver') or '--'))}<br><br>签字/日期：</div></div>
<h2>1. 项目概况</h2>
<p>操作员：{html.escape(str(project.get('operator', '--')))}；设备：{html.escape(str(project.get('device_model', '--')))}；坐标系统：<code>{html.escape(str(project.get('coordinate_system', '--')))}</code>；高程基准：{html.escape(str(project.get('vertical_datum','--')))}。</p>
<h2>2. 测线清单</h2>{table(['line_id','name','length_m','data_quality','rtk_status','sensor_sync_status','processing_status','target_count','data_format'], lines)}
<h2>3. 数据质检</h2>{table(['line_id','line_name','status','sample_count','trace_count','length_m','orientation','orientation_message','issue_count'], quality)}
<h2>4. 雷达—RTK—IMU 同步</h2>{table(['line_id','status','rtk_coverage_ratio','rtk_fixed_ratio','rtk_max_residual_s','imu_coverage_ratio','altimeter_coverage_ratio','gap_count','jump_count','warning_count'], sensor_sync)}
<h2>5. 数据处理与可追溯</h2>{table(['artifact_id','line_id','method_id','method_name','role','status','input_shape','output_shape','output_data_sha256'], artifacts)}
<h2>6. 基覆界面人工标注</h2>{table(['line_id','status','version','keypoint_count','coverage_ratio','judged_ratio','weak_ratio','ignore_ratio','no_interface_ratio','spatial_curve_path'], interfaces)}
<h2>7. 空间成果与 GIS 图层</h2>{table(['line_id','spatial_csv_path','row_count','has_xy_count','empty_xy_count'], spatial)}
{table(['name','kind','role','crs','geometry_type','is_dem','bounds','source_path'], gis_layers)}
<h2>8. 钻孔对比</h2>{table(['borehole_id','line_id','trace_index','borehole_depth_m','interpreted_depth_m','error_m','absolute_error_m','threshold_m','status'], boreholes)}
<h2>9. 工程图件</h2>{''.join(f'<figure><img src="../figures/{html.escape(p.name)}"><figcaption>{html.escape(p.stem)}</figcaption></figure>' for p in figure_paths)}
<h2>附录 A：历史点目标兼容数据</h2>{table(['target_id','line_id','distance_m','depth_m','type','status'], targets)}
<p class="meta">说明：正式成果以连续基覆界面曲线及其弱可见、忽略、无界面语义为准；历史点目标仅作兼容附录。</p>
</body></html>"""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    tmp.write_text(html_text, encoding="utf-8")
    tmp.replace(path)

def _write_pdf_report(path: Path, *, summary: dict[str, Any], lines: list[dict[str, Any]], quality: list[dict[str, Any]], interfaces: list[dict[str, Any]], targets: list[dict[str, Any]], artifacts: list[dict[str, Any]], spatial: list[dict[str, Any]], sensor_sync: list[dict[str, Any]] | None = None, gis_layers: list[dict[str, Any]] | None = None, boreholes: list[dict[str, Any]] | None = None, figure_paths: list[Path] | None = None) -> None:
    """Write the paginated engineering PDF with cover, approvals and appendices.

    CSV/JSON/XLSX remain the normalized evidence tables and the checksum manifest
    provides independent package-integrity verification.
    """
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib import font_manager
    import matplotlib.pyplot as plt

    def _choose_font() -> str:
        candidates = ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Noto Sans CJK JP", "AR PL UMing CN", "AR PL KaitiM GB", "WenQuanYi Micro Hei", "Arial Unicode MS", "DejaVu Sans"]
        installed = {f.name for f in font_manager.fontManager.ttflist}
        for name in candidates:
            if name in installed:
                return name
        return "DejaVu Sans"

    configured_font = configure_matplotlib_cjk_fonts()
    font_name = configured_font if configured_font != "DejaVu Sans" else _choose_font()
    sensor_sync = sensor_sync or []
    gis_layers = gis_layers or []
    boreholes = boreholes or []
    figure_paths = figure_paths or []
    page_number = 0

    def _page(title: str):
        nonlocal page_number
        page_number += 1
        fig = plt.figure(figsize=(8.27, 11.69), dpi=100)
        fig.patch.set_facecolor("white")
        ax = fig.add_axes([0, 0, 1, 1])
        ax.axis("off")
        fig.text(0.06, 0.955, title, fontsize=18, fontweight="bold", fontname=font_name)
        fig.text(0.06, 0.928, "MyGPR 正式工程成果报告 · 可审计交付包", fontsize=9, color="#4b5563", fontname=font_name)
        fig.text(0.50, 0.025, f"第 {page_number} 页", fontsize=8, color="#64748B", ha="center", fontname=font_name)
        fig.text(0.94, 0.025, str(summary.get("report_revision", "R1")), fontsize=8, color="#64748B", ha="right", fontname=font_name)
        return fig

    def _kv_lines(fig, start_y: float, rows: list[tuple[str, Any]]) -> float:
        y = start_y
        for key, value in rows:
            fig.text(0.075, y, str(key), fontsize=10, fontweight="bold", fontname=font_name)
            fig.text(0.24, y, str(value), fontsize=10, fontname=font_name)
            y -= 0.027
        return y

    def _table_page(pdf: PdfPages, title: str, headers: list[str], rows: list[dict[str, Any]], limit: int = 24) -> None:
        fig = _page(title)
        ax = fig.add_axes([0.05, 0.08, 0.90, 0.80])
        ax.axis("off")
        data = [[str(row.get(h, ""))[:42] for h in headers] for row in rows[:limit]]
        if not data:
            data = [["无记录"] + [""] * (len(headers) - 1)]
        table = ax.table(cellText=data, colLabels=headers, loc="upper left", cellLoc="left")
        table.auto_set_font_size(False)
        table.set_fontsize(7)
        table.scale(1.0, 1.35)
        for (_row, _col), cell in table.get_celld().items():
            cell.set_edgecolor("#d1d5db")
            cell.set_linewidth(0.4)
            cell.get_text().set_fontname(font_name)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    project = summary.get("project", {})
    metrics = summary.get("metrics", {})
    with PdfPages(tmp) as pdf:
        fig = _page(f"{project.get('name', 'MyGPR 项目')} 成果报告")
        y = _kv_lines(
            fig,
            0.86,
            [
                ("生成时间", summary.get("generated_at", "")),
                ("项目编号", project.get("project_no", "--")),
                ("测区位置", project.get("location", "--")),
                ("操作员", project.get("operator", "--")),
                ("设备型号", project.get("device_model", "--")),
                ("坐标系统", project.get("coordinate_system", "--")),
                ("高程基准", project.get("vertical_datum", "--")),
                ("项目快照", (summary.get("snapshot") or {}).get("snapshot_id", "--")),
                ("空间成果版本", (summary.get("source_binding") or {}).get("spatial_result_id", "未绑定") or "未绑定"),
                ("项目路径", project.get("project_path", "--")),
            ],
        )
        metric_rows = [
            ("测线数量", metrics.get("line_count", 0)),
            ("已导入测线", metrics.get("imported_line_count", 0)),
            ("质检通过", metrics.get("qc_passed_count", 0)),
            ("质检警告", metrics.get("qc_warning_count", 0)),
            ("处理结果", metrics.get("processing_artifact_count", 0)),
            ("界面标注测线", metrics.get("interface_line_count", 0)),
            ("空间成果", metrics.get("spatial_export_count", 0)),
            ("钻孔误差达标", f"{metrics.get('borehole_passed_count', 0)}/{metrics.get('borehole_comparison_count', 0)}"),
        ]
        fig.text(0.075, y - 0.035, "核心统计", fontsize=13, fontweight="bold", fontname=font_name)
        _kv_lines(fig, y - 0.075, metric_rows)
        fig.text(0.075, 0.20, "交付说明", fontsize=13, fontweight="bold", fontname=font_name)
        fig.text(0.075, 0.17, "PDF 为交付摘要；CSV、JSON、HTML 文件保留完整可审计数据。", fontsize=10, fontname=font_name)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        fig = _page("审批与签署")
        approval = [
            ("编制", project.get("compiler") or project.get("operator") or "--"),
            ("复核", project.get("reviewer") or "--"),
            ("批准", project.get("approver") or "--"),
        ]
        y = 0.82
        for role, person in approval:
            fig.text(0.10, y, role, fontsize=12, fontweight="bold", fontname=font_name)
            fig.text(0.25, y, str(person), fontsize=11, fontname=font_name)
            fig.text(0.55, y, "签字：________________", fontsize=10, fontname=font_name)
            fig.text(0.76, y, "日期：____________", fontsize=10, fontname=font_name)
            y -= 0.16
        fig.text(0.10, 0.28, "报告版本", fontsize=11, fontweight="bold", fontname=font_name)
        fig.text(0.25, 0.28, str(summary.get("report_revision", "R1")), fontsize=11, fontname=font_name)
        fig.text(0.10, 0.22, "数据完整性", fontsize=11, fontweight="bold", fontname=font_name)
        fig.text(0.25, 0.22, "详见 checksums.sha256 与 file_audit.csv", fontsize=10, fontname=font_name)
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        fig = _page("目录")
        chapters = ["1 项目概况", "2 测线与数据质检", "3 雷达—RTK—IMU 同步", "4 数据处理与可追溯", "5 基覆界面标注", "6 空间成果与 GIS", "7 钻孔对比", "8 工程图件", "附录 数据清单与校验"]
        y=0.84
        for idx, chapter in enumerate(chapters, start=1):
            fig.text(0.12,y,chapter,fontsize=11,fontname=font_name)
            fig.text(0.88,y,str(idx+3),fontsize=10,ha="right",fontname=font_name)
            y-=0.065
        pdf.savefig(fig,bbox_inches="tight"); plt.close(fig)

        _table_page(pdf, "测线清单", ["line_id", "name", "length_m", "data_quality", "rtk_status", "sensor_sync_status", "processing_status", "target_count"], lines)
        _table_page(pdf, "数据质检", ["line_id", "line_name", "status", "sample_count", "trace_count", "orientation_message"], quality)
        _table_page(pdf, "雷达—RTK—IMU 同步", ["line_id", "status", "rtk_coverage_ratio", "rtk_fixed_ratio", "rtk_max_residual_s", "imu_coverage_ratio", "gap_count", "jump_count", "warning_count"], sensor_sync)
        _table_page(pdf, "处理结果与哈希", ["artifact_id", "line_id", "method_id", "role", "status", "output_shape", "output_data_sha256"], artifacts)
        _table_page(pdf, "基覆界面标注", ["line_id", "status", "keypoint_count", "coverage_ratio", "judged_ratio", "weak_ratio", "ignore_ratio", "no_interface_ratio"], interfaces)
        _table_page(pdf, "空间成果", ["line_id", "spatial_csv_path", "row_count", "has_xy_count", "empty_xy_count"], spatial)
        _table_page(pdf, "GIS 图层", ["name", "kind", "role", "crs", "geometry_type", "is_dem", "source_path"], gis_layers)
        _table_page(pdf, "钻孔对比", ["borehole_id", "line_id", "trace_index", "borehole_depth_m", "interpreted_depth_m", "error_m", "absolute_error_m", "threshold_m", "status"], boreholes)
        for figure_path in figure_paths:
            if not figure_path.exists():
                continue
            fig = _page(f"工程图件 · {figure_path.stem}")
            image = plt.imread(figure_path)
            ax = fig.add_axes([0.06, 0.10, 0.88, 0.78]); ax.imshow(image); ax.axis("off")
            pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)
    tmp.replace(path)

__all__ = ['_write_xlsx_report', '_write_report_figures', '_write_checksums', '_write_html_report', '_write_pdf_report']
