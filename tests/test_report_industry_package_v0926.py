from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

from core.field_project_operations import import_line_data
from core.field_project_store import FieldProjectStore
from core.field_report_export import generate_project_report_package


def test_formal_report_contains_excel_audit_figures_approvals_and_checksums(tmp_path: Path) -> None:
    store = FieldProjectStore.create_empty(
        tmp_path / "project",
        name="野外基覆界面勘探",
        project_no="MYGPR-2026-001",
        coordinate_system="EPSG:32648",
        vertical_datum="1985国家高程基准",
    )
    import_line_data(
        store,
        Path("sample_data/gui_sidecar_all_data_main.csv"),
        line_id="L01",
        name="一号测线",
    )
    result = generate_project_report_package(
        store,
        package_name="formal_delivery",
        report_profile={
            "compiler": "编制人",
            "reviewer": "复核人",
            "approver": "批准人",
            "revision": "A",
        },
    )

    package = store.root / result.package_dir
    xlsx = store.root / result.xlsx_path
    audit = store.root / result.audit_csv_path
    checksums = store.root / result.checksums_path
    figures = store.root / result.figures_dir
    pdf = store.root / result.pdf_path
    html = store.root / result.html_path

    assert xlsx.exists() and xlsx.stat().st_size > 0
    assert audit.exists() and checksums.exists()
    assert pdf.exists() and pdf.stat().st_size > 0
    assert html.exists() and "批准人" in html.read_text(encoding="utf-8")
    assert figures.is_dir() and any(figures.glob("*.png"))

    workbook = load_workbook(xlsx, read_only=True)
    assert {"测线清单", "数据质检", "传感器同步", "基覆界面", "处理记录", "空间成果", "GIS图层"}.issubset(workbook.sheetnames)
    workbook.close()

    with audit.open("r", encoding="utf-8-sig", newline="") as handle:
        audit_rows = list(csv.DictReader(handle))
    assert any(row["path"] == "project_report.pdf" for row in audit_rows)
    assert all(len(row["sha256"]) == 64 for row in audit_rows)

    checksum_entries = {}
    for line in checksums.read_text(encoding="utf-8").splitlines():
        digest, rel = line.split("  ", 1)
        checksum_entries[rel] = digest
    assert "project_report.pdf" in checksum_entries
    actual_pdf_hash = hashlib.sha256((package / "project_report.pdf").read_bytes()).hexdigest()
    assert checksum_entries["project_report.pdf"] == actual_pdf_hash

    manifest = json.loads((store.root / result.manifest_path).read_text(encoding="utf-8"))
    assert manifest["summary"]["report_revision"] == "A"
    assert manifest["summary"]["approval"]["compiler"] == "编制人"
